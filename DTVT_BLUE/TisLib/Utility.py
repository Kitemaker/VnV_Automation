import sys
import os.path
import logging
import xml.etree.ElementTree as ET
from openpyxl import workbook, worksheet, cell
from openpyxl import load_workbook


def Arrange(DeviceID:list,Kp:list,rev:bool):

    """Arrange the DeviceID list in increasing or decreasing values of Kp List.
       DeviceID= List of Device to be arranged,Kp =Kp Values
       rev= True for Increasing,False for decreasing
    """
   
    _ids = list(DeviceID)
    _kps = list(Kp)
    outids = []
    try:
        if len(_ids) == len(_kps):
            _kps.sort(reverse=rev)
            for  i in range(len(_kps)):
                outids.append(_ids[Kp.index(_kps[i])])        
        else:
            return []                
    except:
       logging.error("Unexpected error:", sys.exc_info()[0])
    finally:        
        return outids

def getSDDBBetSignals(KpSig1:int,KpSig2:int,Track:str,SDDBData:dict):   
    count=0
    try:
        sddbKps = GetKpValue(SDDBData['KpValue'],SDDBData['KpCorrected_Trolley_Value'])
        sddbTrack = SDDBData['Track_ID']
        sddbTrackKps = []
        for item in range(len(sddbTrack)):
            if(sddbTrack[item] == Track):
                sddbTrackKps.append(sddbKps[item])

        for k in range(len(sddbTrackKps)):
            if(KpSig1<KpSig2):
                if(int(KpSig1)<int(sddbTrackKps[k]) & int(sddbTrackKps[k]) < int(KpSig2)):
                    count=count+1
            elif(KpSig1>KpSig2):
                if(int(KpSig1)>int(sddbTrackKps[k]) & int(sddbTrackKps[k]) > int(KpSig2)):
                    count=count+1
    except:
        logging.error("Error:Fun = getSDDBBetSignals,Module=Utility", sys.exc_info()[0])
    finally:        
        return count

def GetLogger(logFile):
    
    logger = logging.getLogger('Tis_Logger')
    logger.setLevel(logging.DEBUG)
    # create file handler which logs even debug messages
    fh = logging.FileHandler(logFile,mode='w')
    fh.setLevel(logging.DEBUG)
    # create console handler with a higher log level
    ch = logging.StreamHandler()
    ch.setLevel(logging.ERROR)
    # create formatter and add it to the handlers
    formatter = logging.Formatter('%(asctime)s : %(levelname)s : %(message)s',datefmt='%Y/%m/%d %I:%M:%S %p')
    ch.setFormatter(formatter)
    fh.setFormatter(formatter)
    # add the handlers to logger
    logger.addHandler(ch)
    logger.addHandler(fh)

    return logger

def GetProjectConstants(configfile:str):
    """Returns the dictionalry of project constants defined in the constant file
       constantfile= full path of constant file
    """
    tis_logger=logging.getLogger('Tis_Logger')
    proj_const=dict()
    try:    
        config = ET.parse(configfile)
        rootXml = config.getroot()
        const_file_node = rootXml.find("ConstantFile")
        const_file = ET.parse(const_file_node.attrib['path'])   
        const_root_xml = const_file.getroot()

        for el in const_root_xml.findall('constant'):
            # return the value 'None' if attrib is not found
            try:               
                proj_const[el.get('name')] = el.get('value')         
            except:
                tis_logger.error("Method=GetProjectConstants , Module=Utility" +  sys.exc_info()[0])
                continue
    except:
       tis_logger.error("Method=GetProjectConstants , Module=Utility" ,  sys.exc_info()[0])
    finally:
        return proj_const

def GetSDDofPoint(point,SDD_Cap):
    for item in range(len(SDD_Cap['Name'])):
        sdd_name=''
        pt_list  = SDD_Cap['Point_ID_List'][item].split(';')
        if point in pt_list:
            sdd_name = SDD_Cap['Name'][item]
            return sdd_name

def GetPointsOfSwitch(switchIndex, SwitchsCap):
    try:

        convergent_Points = SwitchsCap['Convergent_Point_ID_List'][switchIndex] 
        divergent_Points  = SwitchsCap['Divergent_Point_ID_List'][switchIndex]
        if convergent_Points == '0' and divergent_Points == '0':
            return []
        elif convergent_Points == '0' and divergent_Points != '0':
            return divergent_Points.split(';')
        elif convergent_Points != '0' and divergent_Points == '0':
            return convergent_Points.split(';')
        else:
            return convergent_Points.split(';') + divergent_Points.split(';')
    except:
        tis_logger.error("Method = GetPointsOfSwitch , Module = Utility" +  sys.exc_info()[0])
        return []

def ConvertToString(listInput:list):
    out_text=''
    for item in listInput:
        out_text=out_text + item +';'

    return out_text

def GetKpValue(kpValue:list,kpTrolleyValue:list):
    kp=list()
    try:    
        if (len(kpValue) == len(kpTrolleyValue)):
            for i in range(len(kpValue)):
                kp.append(int(kpValue[i]) + int(kpTrolleyValue[i]))

            return kp
               
        else:
            return None
        
    except:
        print("Unexpected error:" + sys.exc_info()[0])
        return None

def Get_C_Late_Change_Distance(InputFile:str):
    """ Input File must be in .xlsx format and Input must be in 'Sheet1' 
        returns dictionary of platform name and its C_Late_Change_Distance
    """
    c_late_change_dist = dict()
    try:
        
        input_wb = load_workbook(InputFile)
        input_ws = input_wb.get_sheet_by_name('Sheet1')
        col_A = input_ws['A']
        col_B = input_ws['B']
        for index in range(1,len(col_A)):
            if col_B[index].value != None:
                c_late_change_dist[str.strip(col_A[index].value)] = int(col_B[index].value)
            else:
                break        
    except:
        logging.getLogger('Tis_Logger').error(sys.exc_info()[0])
        c_late_change_dist = None
    finally:
        input_wb.close()
        return c_late_change_dist 

def GetRootFolder():
    return os.path.abspath(os.path.join(os.path.dirname(__file__),".."))

def GetSSPType(SSP:str,PlatformsSSP:list,StablingSSP:list):
    """
     find is SSP is associated with Platform or Stabling or Others
     Input: SSP Name , Column 'SSP_ID_List' from Platforms_Cap and Stablings_Location_Cap
     Return values :  "Others", "Platform","Stabling"
    """
    _sspType = "Others"
    for plt in PlatformsSSP:
        if SSP in plt.split(';'):
            _sspType = "Platform"
            return _sspType
    for sta in StablingSSP:
        if SSP in sta.split(';'):
            _sspType = "Stabling"
            return _sspType

    return _sspType

def GetTrackDirections(InputFile:str):

    """ Input File must be in .xlsx format and Input must be in 'Sheet1' 
        returns dictionary of track name and its direction in increasing Kp
    """
    track_directions = dict()
    try:
        
        input_wb = load_workbook(InputFile)
        input_ws = input_wb.get_sheet_by_name('Sheet1')
        col_A = input_ws['A']
        col_B = input_ws['B']
        for index in range(1,len(col_A)):
            if col_B[index].value != None:
                track_directions[str.strip(col_A[index].value)] = col_B[index].value
            else:
                break        
    except:
        logging.getLogger('Tis_Logger').error(sys.exc_info()[0])
        track_directions = None
    finally:
        input_wb.close()
        return track_directions 
 
    
def GetNearestTailSDDBForSSP(SSPKp:int,Dir:str,SddbKp:list):
    """
    Returns the Kp value of Nearest SDDB in tail direction for the given SSP
    Input SSP Kp value, SSP Direction, SDDB Kp List
    """   
    sddb_kp = 0
    try:

        if str.upper(Dir) == 'UP':
            SddbKp.sort(reverse = True)
            for index in range(len(SddbKp)):
                if SSPKp > SddbKp[index] :
                    sddb_kp = SddbKp[index]
                    break

        if str.upper(Dir) == 'DOWN':
            SddbKp.sort()
            for index in range(len(SddbKp)):
                if SddbKp[index] > SSPKp:
                    sddb_kp = SddbKp[index]
                    break

    except:
        logging.getLogger('Tis_Logger').error(sys.exc_info()[0])
        sddb_kp = None
    finally:        
        return sddb_kp

def GetTrackDirectionFile(RootFolder:str):
    try:
        return os.path.join(RootFolder,"Input\\Constant\\Tracks_Direction.xlsx")
    except:
        logging.getLogger('Tis_Logger').error("Unexpected Error in Module: Utility.py, Method;GetTrackDirectionFile", sys.exc_info()[0])
        return None



def WriteToWorkSheet(Worksheet:worksheet, RowIndex:int, ColumnValues:list):
    try:

        for item in range(len(ColumnValues)):
            Worksheet.cell(row = RowIndex, column = item + 1, value =  ColumnValues[item])
    except:
        logging.getLogger('Tis_Logger').error("Unexpected Error in Module: Utility.py, Method;WriteToWorkSheet", sys.exc_info()[0])
   

def SaveReport(Workbook:workbook,GlobalTestResults:list,RootFolder,ResultFile):
       
    try:
        if "NOK" in GlobalTestResults:
            result_file_path = os.path.join(RootFolder + "\\Results\\" ,'Test_Results_NOK_' + ResultFile)
        else:
            result_file_path = os.path.join(RootFolder + "\\Results\\" ,'Test_Results_OK_' + ResultFile)

        Workbook.save(result_file_path)
        logging.getLogger('Tis_Logger').info("Report generated successfully at " + result_file_path ) 
        print("Report generated successfully at " + result_file_path )
        Workbook.close()
    except:
        logging.getLogger('Tis_Logger').error("Unexpected error in writing report :" , sys.exc_info()[0])

