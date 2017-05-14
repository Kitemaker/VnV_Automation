#========================Template ============================================================================================
import sys
import os,os.path
import logging
from openpyxl import workbook, worksheet

# add root folder in search path
root_folder = os.path.abspath(os.path.join(os.path.dirname(__file__),".."))
sys.path.append(root_folder)

import TisLib
from TisLib import Utility, Configuration , CSVReader

try:
    # Get Tis_Logger and specify the log file name   
    log_file_path = os.path.join(root_folder + "\\Logs\\" , os.path.basename(__file__).split('.')[0] + '.log')
    result_file_name = os.path.basename(__file__).split('.')[0] + '.xlsx'
    result_file_path = os.path.join(root_folder + "\\Results\\" ,'Test_Results_' + result_file_name)
    tis_log = Utility.GetLogger(log_file_path)
except:
    print("Error while creating logger")

config = Configuration.Configuration(root_folder)
# read constant
proj_const = dict()
proj_const = Utility.GetProjectConstants(config.config_File_Path)
print("Executing " + os.path.basename(__file__))
tis_log.info("Executing " + os.path.basename(__file__))
### ======== Template Ends=======================================================================================================

csv_reader = CSVReader.CSVReader(config.csv_folder_path)
SyDT = csv_reader.SyDT  
signal_cap = SyDT[csv_reader.Signals_Cap]
sddb_cap = SyDT[csv_reader.SDDB_Cap]
tracks_cap =SyDT[csv_reader.Tracks_Cap]
signals_id = signal_cap['ID']
signal_name = signal_cap['Name']
signal_kp = Utility.GetKpValue(signal_cap['KpValue'] ,signal_cap['KpCorrected_Trolley_Value'])
signal_track = signal_cap['Track_ID']
signal_dir = signal_cap['Direction']
Signal_Type_Function = signal_cap['Signal_Type_Function']

sigUp,sigUpKp,sigDn,sigDnKp = [],[],[],[]
sigIdSortedUp,sigIdSortedDn = [],[]
# Create report file
global_test_results = list()
wbReport = workbook.Workbook()
wsRpt=wbReport.active
wsRpt.title = "Rule_TMS_SDD_0013"
rwCount = 1
wsRpt.cell(row = rwCount, column = 1, value = "Sig1 ID")
wsRpt.cell(row = rwCount, column = 2, value = "Sig1_Type_Function")
wsRpt.cell(row = rwCount, column = 3, value = "Sig1 Kp")
wsRpt.cell(row = rwCount, column = 4, value = "Sig1 Track")
wsRpt.cell(row = rwCount, column = 5, value = "Sig2 ID")
wsRpt.cell(row = rwCount, column = 6, value = "Sig2_Type_Function")
wsRpt.cell(row = rwCount, column = 7, value = "Sig2 Kp")
wsRpt.cell(row = rwCount, column = 8, value = "Sig2 Track")
wsRpt.cell(row = rwCount, column = 9, value = "Dir")
wsRpt.cell(row = rwCount, column = 10, value = "SDDB Count")

rwCount = rwCount+1

for tr in tracks_cap['Name']:

    sigUp, sigUpKp , sigDn , sigDnKp= [],[],[],[]      
    tis_log.info("Getting Signals for track : " + tr)

    for item in range(len(signal_track)):
        if(tr == signal_track[item]) and (signal_dir[item] == 'Up') and (Signal_Type_Function[item] != 'Virtual'):
            sigUp.append(signals_id[item])
            sigUpKp.append(signal_kp[item])
        if (tr == signal_track[item]) and (signal_dir[item] == 'Down') and (Signal_Type_Function[item] != 'Virtual'):
            sigDn.append(signals_id[item])
            sigDnKp.append(signal_kp[item])                
    
    sigIdSortedUp = Utility.Arrange(sigUp,sigUpKp,False)  
          
    for i in range(len(sigIdSortedUp)-1):
        kp1 = signal_kp[signals_id.index(sigIdSortedUp[i])]
        kp2 = signal_kp[signals_id.index(sigIdSortedUp[i+1])]
        sddbCount = Utility.getSDDBBetSignals(kp1,kp2,tr,sddb_cap)
                                      
        wsRpt.cell(row = rwCount, column = 1, value = signal_name[signals_id.index(sigIdSortedUp[i])])
        wsRpt.cell(row = rwCount, column = 2, value = Signal_Type_Function[signals_id.index(sigIdSortedUp[i])])
        wsRpt.cell(row = rwCount, column = 3, value = str(kp1))
        wsRpt.cell(row = rwCount, column = 4, value = signal_track[signals_id.index(sigIdSortedUp[i])])
        wsRpt.cell(row = rwCount, column = 5, value = signal_name[signals_id.index(sigIdSortedUp[i+1])])
        wsRpt.cell(row = rwCount, column = 6, value = Signal_Type_Function[signals_id.index(sigIdSortedUp[i+1])])
        wsRpt.cell(row = rwCount, column = 7, value = str(kp2))
        wsRpt.cell(row = rwCount, column = 8, value = signal_track[signals_id.index(sigIdSortedUp[i+1])])
        wsRpt.cell(row = rwCount, column = 9, value = "Up")
        wsRpt.cell(row = rwCount, column = 10, value = str(sddbCount))

        if(sddbCount == 0):           
            tis_log.error("SDDB Count =0 for  " + signal_name[signals_id.index(sigIdSortedUp[i])]+"," + signal_name[signals_id.index(sigIdSortedUp[i+1])])
        rwCount=rwCount+1
            
    
    sigIdSortedDn=Utility.Arrange(sigDn,sigDnKp,False)
    for i in range(len(sigDnKp)-1):
        wsRpt.cell(row = rwCount, column = 1, value = signal_name[signals_id.index(sigIdSortedDn[i])])
        wsRpt.cell(row = rwCount, column = 2, value = Signal_Type_Function[signals_id.index(sigIdSortedDn[i])])
        wsRpt.cell(row = rwCount, column = 3, value = str(kp1))
        wsRpt.cell(row = rwCount, column = 4, value = signal_track[signals_id.index(sigIdSortedDn[i])])
        wsRpt.cell(row = rwCount, column = 5, value = signal_name[signals_id.index(sigIdSortedDn[i+1])])
        wsRpt.cell(row = rwCount, column = 6, value = Signal_Type_Function[signals_id.index(sigIdSortedDn[i+1])])
        wsRpt.cell(row = rwCount, column = 7, value = str(kp2))
        wsRpt.cell(row = rwCount, column = 8, value = signal_track[signals_id.index(sigIdSortedDn[i+1])])
        wsRpt.cell(row = rwCount, column = 9, value = "Dn")
        wsRpt.cell(row = rwCount, column = 10, value = str(sddbCount))
        if(sddbCount==0):            
            tis_log.error("SDDB Count =0 for  " + signal_name[signals_id.index(sigIdSortedDn[i])]+" , " + signal_name[signals_id.index(sigIdSortedDn[i+1])])
            global_test_results.append('NOK')
        else:
            global_test_results.append('OK')
        rwCount=rwCount+1
Utility.SaveReport(wbReport,global_test_results,root_folder,result_file_name)

