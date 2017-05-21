### ========= Template =========================================================================
import sys , os , os.path
import logging
from openpyxl import workbook, worksheet
# add root folder in search path
root_folder = os.path.abspath(os.path.join(os.path.dirname(__file__),".."))
sys.path.append(root_folder)
import DtvtLib
from DtvtLib import Utility, Configuration , CSVReader
param = Utility.Initialise(root_folder,__file__)
log_file_path    = param['log_file_path']
result_file_name = param['result_file_name']
result_file_path = param['result_file_path']
dtvt_log         = param['dtvt_log']
proj_const       = param['proj_const']
SyDT             = param['SyDT']

### ======== Template Ends===========================================================================

### =========Rule RSCMA_SPECIFIC_RMZ_CONSTRAINT ===================================================================
# Author : Saleem Javed
# Updated 11 April 2017
# Caps Used : Reverse_Movement_Zones_Cap  , Points_Cap
# Constant used :  Min_RMR_Protection ,  Min_RMR_Protection_Forward
###==================================================================================================

# Read Caps
rmz_cap = SyDT['Reverse_Movement_Zones_Cap']
point_cap = SyDT['Points_Cap']

rmz_id = rmz_cap['ID']
rmz_names = rmz_cap['Name']
rmz_track = rmz_cap['Track_ID']
rmz_dir = rmz_cap['Reverse_Movement_Direction']
rmz_type = rmz_cap['RMZ_Type']
rmz_kp_begin = Utility.GetKpValue(rmz_cap['Kp_BeginValue'],rmz_cap['Kp_BeginCorrected_Trolley_Value'])
rmz_kp_end =   Utility.GetKpValue(rmz_cap['Kp_EndValue'],rmz_cap['Kp_EndCorrected_Trolley_Value'])

point_name = point_cap['Name']
point_track  = point_cap['Track_ID']
point_kp_toe = Utility.GetKpValue(point_cap['Kp_ToeValue'],point_cap['Kp_ToeCorrected_Trolley_Value'])
point_kp_fouling = Utility.GetKpValue(point_cap['Kp_Fouling_PointValue'],point_cap['Kp_Fouling_PointCorrected_Trolley_Value'])

Min_RMR_Protection = int(proj_const['Min_RMR_Protection'])
Min_RMR_Protection_Forward = int(proj_const['Min_RMR_Protection_Forward'])

# Create report file
global_test_results = list()
wbReport = workbook.Workbook()
wsRpt=wbReport.active
wsRpt.title = "Test_Results"
rwCount = 1

Utility.WriteToWorkSheet(wsRpt,rwCount,["ID", "RMZ", "Direction","RMZ_Type",
                                        "Track_ID","Kp_Begin_Value","Kp_End_Value",
                                        "Min RMR Protection UpExtr 1","Min RMR Protection UpExtr 2","Min RMR Protection Dn Extr 1", "Min RMR Protection Dn Extr 2",
                                        "Point ID in Protection Zone", "Fouling Point ID in Protection Zone",
                                        "Result"])


for index in range(len(rmz_names)):
    if rmz_type[index] == 'RMR_Per_Zones':
        rwCount = rwCount+1  
        try:
            _result= []
            _rmz = rmz_names[index].strip()
            _rmz_kp_b = rmz_kp_begin[index]
            _rmz_kp_e = rmz_kp_end[index]
            _rmz_type = rmz_type[index]
            _rmz_track = rmz_track[index]
            if _rmz_kp_b < _rmz_kp_e:
                #Increasing Kp in conventional direction
                DnK1 = _rmz_kp_b - Min_RMR_Protection
                DnK2 = _rmz_kp_b
                UpK1 = _rmz_kp_e + Min_RMR_Protection
                UpK2 = _rmz_kp_e  
                
            elif  _rmz_kp_b > _rmz_kp_e:
                # Decreaing Kp in conventional direction
                DnK1 = _rmz_kp_b + Min_RMR_Protection
                DnK2 = _rmz_kp_b
                UpK1 = _rmz_kp_e - Min_RMR_Protection
                UpK2 = _rmz_kp_e 
                                         
            _pt_toe_fail_list = ''
            _pt_fpn_fail_list = ''
            for item in range(len(point_name)):
                _pt_name = point_name[item]
                _pt_toe = point_kp_toe[item]
                _pt_flng =point_kp_fouling[item]
                _condition1 = None
                _condition2 = None
                                
                if point_track[item] == _rmz_track and rmz_dir[index] == "Both":
                    _condition1 = (( _pt_toe < UpK1  and _pt_toe < UpK2 )    or ( _pt_toe > UpK1  and _pt_toe > UpK2 )) and  (( _pt_toe < DnK1  and  _pt_toe < DnK2 )   or ( _pt_toe > DnK1  and  _pt_toe > DnK2 ))
                    _condition2 = (( _pt_flng < UpK1 and _pt_flng < UpK2 )   or ( _pt_flng > UpK1 and _pt_flng > UpK2 )) and  (( _pt_flng < DnK1 and  _pt_flng < DnK2 )  or ( _pt_flng > DnK1 and  _pt_flng > DnK2 ))              
                    
                elif point_track[item] == _rmz_track and rmz_dir[index] == "Up":
                    _condition1 = ( _pt_toe < UpK1  and _pt_toe < UpK2 )   or ( _pt_toe > UpK1  and _pt_toe > UpK2 )
                    _condition2 = ( _pt_flng < UpK1 and _pt_flng < UpK2 )  or ( _pt_flng > UpK1 and _pt_flng > UpK2 ) 

                elif point_track[item] == _rmz_track and rmz_dir[index] == "Down":
                    _condition1 = ( _pt_toe < DnK1   and  _pt_toe < DnK2 )   or  ( _pt_toe > DnK1   and  _pt_toe > DnK2 )
                    _condition2 = ( _pt_flng < DnK1  and  _pt_flng < DnK2 )  or  ( _pt_flng > DnK1  and  _pt_flng > DnK2 )                

                if( _condition1 == True and _condition2 == True):                    
                    _result.append('OK')
                    global_test_results.append('OK')

                elif( _condition1 == False and _condition2 == True):
                    _result.append('NOK')
                    global_test_results.append('NOK')
                    _pt_toe_fail_list = _pt_toe_fail_list + ','+ _pt_name

                elif(_condition1 == True and _condition2 == False):
                    _result.append('NOK')
                    global_test_results.append('NOK')   
                    _pt_fpn_fail_list = _pt_fpn_fail_list + ','+ _pt_name
                elif(_condition1 == False and _condition2 == False):
                    global_test_results.append('NOK')
                    _result.append('NOK')
                    _pt_toe_fail_list = _pt_toe_fail_list + ','+ _pt_name   
                    _pt_fpn_fail_list = _pt_fpn_fail_list + ','+ _pt_name  


            Utility.WriteToWorkSheet(wsRpt, rwCount, [ rmz_id[index], _rmz,rmz_dir[index] , _rmz_type, _rmz_track, _rmz_kp_b, _rmz_kp_e, UpK1, UpK2, DnK1, DnK2,_pt_toe_fail_list, _pt_fpn_fail_list])
            if 'NOK' in _result:
                wsRpt.cell(row = rwCount, column = 14, value = 'NOK')
                dtvt_log.error('RMZ:' + _rmz + ', Track: '+ _rmz_track + ' has failed the rule for Points: ' + _pt_toe_fail_list +'\t' + _pt_fpn_fail_list)
                
            else:
                wsRpt.cell(row = rwCount, column = 14, value = 'OK')

        except:
            dtvt_log.error( "Module:RULE_SCMA_SPECIFIC_RMZ_CONSTRAINT.py" + "Method:__main__ " +"item =" + rmz_names[index]  + "\t" ,  sys.exc_info()[0])               
            
for index in range(len(rmz_names)):
    if rmz_type[index] == 'RMR_Per_Zones':
        rwCount = rwCount+1
        try:
            _result= []
            _rmz = rmz_names[index].strip()
            _rmz_kp_b = rmz_kp_begin[index]
            _rmz_kp_e = rmz_kp_end[index]
            _rmz_type = rmz_type[index]
            _rmz_track = rmz_track[index]
            if _rmz_kp_b < _rmz_kp_e:
                #Increasing Kp in conventional direction
                DnK1 = _rmz_kp_e + Min_RMR_Protection_Forward
                DnK2 = _rmz_kp_e
                UpK1 = _rmz_kp_b - Min_RMR_Protection_Forward
                UpK2 = _rmz_kp_b  
                
            elif  _rmz_kp_b > _rmz_kp_e:
                # Decreaing Kp in conventional direction
                DnK1 = _rmz_kp_e - Min_RMR_Protection_Forward
                DnK2 = _rmz_kp_e
                UpK1 = _rmz_kp_b +  Min_RMR_Protection_Forward
                UpK2 = _rmz_kp_b 
           
            _pt_toe_fail_list = ''
            _pt_fpn_fail_list = ''
            for item in range(len(point_name)):
                _pt_name = point_name[item]
                _pt_toe = point_kp_toe[item]
                _pt_flng =point_kp_fouling[item]
                _condition1 = None
                _condition2 = None
                                
                if point_track[item] == _rmz_track and rmz_dir[index] == "Both":
                    _condition1 = (( _pt_toe < UpK1  and _pt_toe < UpK2 )    or ( _pt_toe > UpK1  and _pt_toe > UpK2 )) and  (( _pt_toe < DnK1  and  _pt_toe < DnK2 )   or ( _pt_toe > DnK1  and  _pt_toe > DnK2 ))
                    _condition2 = (( _pt_flng < UpK1 and _pt_flng < UpK2 )   or ( _pt_flng > UpK1 and _pt_flng > UpK2 )) and  (( _pt_flng < DnK1 and  _pt_flng < DnK2 )  or ( _pt_flng > DnK1 and  _pt_flng > DnK2 ))              
                    
                elif point_track[item] == _rmz_track and rmz_dir[index] == "Up":
                    _condition1 = ( _pt_toe < UpK1  and _pt_toe < UpK2 )   or ( _pt_toe > UpK1  and _pt_toe > UpK2 )
                    _condition2 = ( _pt_flng < UpK1 and _pt_flng < UpK2 )  or ( _pt_flng > UpK1 and _pt_flng > UpK2 ) 

                elif point_track[item] == _rmz_track and rmz_dir[index] == "Down":
                    _condition1 = ( _pt_toe < DnK1   and  _pt_toe < DnK2 )   or  ( _pt_toe > DnK1   and  _pt_toe > DnK2 )
                    _condition2 = ( _pt_flng < DnK1  and  _pt_flng < DnK2 )  or  ( _pt_flng > DnK1  and  _pt_flng > DnK2 )                

                if( _condition1 == True and _condition2 == True):                    
                    _result.append('OK')
                    global_test_results.append('OK')

                elif( _condition1 == False and _condition2 == True):
                    _result.append('NOK')
                    global_test_results.append('NOK')
                    _pt_toe_fail_list = _pt_toe_fail_list + ','+ _pt_name

                elif(_condition1 == True and _condition2 == False):
                    _result.append('NOK')
                    global_test_results.append('NOK')   
                    _pt_fpn_fail_list = _pt_fpn_fail_list + ','+ _pt_name
                elif(_condition1 == False and _condition2 == False):
                    global_test_results.append('NOK')
                    _result.append('NOK')
                    _pt_toe_fail_list = _pt_toe_fail_list + ','+ _pt_name   
                    _pt_fpn_fail_list = _pt_fpn_fail_list + ','+ _pt_name  


            Utility.WriteToWorkSheet(wsRpt, rwCount, [ rmz_id[index], _rmz,rmz_dir[index] , _rmz_type, _rmz_track, _rmz_kp_b, _rmz_kp_e, UpK1, UpK2, DnK1, DnK2,_pt_toe_fail_list, _pt_fpn_fail_list])
            if 'NOK' in _result:
                wsRpt.cell(row = rwCount, column = 14, value = 'NOK')
                dtvt_log.error('RMZ:' + _rmz + ', Track: '+ _rmz_track + ' has failed the rule for Points: ' + _pt_toe_fail_list +'\t' + _pt_fpn_fail_list)
                
            else:
                wsRpt.cell(row = rwCount, column = 14, value = 'OK')

        except:
            dtvt_log.error( "Module:RULE_SCMA_SPECIFIC_RMZ_CONSTRAINT.py" + "Method:__main__ " +"item =" + rmz_names[index]  + "\t" ,  sys.exc_info()[0])          

# Save Report
print('Rule Execution Finished. Saveing Report...')
Utility.SaveReport(wbReport,global_test_results,root_folder,result_file_name)

