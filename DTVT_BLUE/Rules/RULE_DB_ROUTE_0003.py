### ========= Template =========================================================================
import sys , os , os.path
import logging
from openpyxl import workbook, worksheet
# add root folder in search path
root_folder = os.path.abspath(os.path.join(os.path.dirname(__file__),".."))
sys.path.append(root_folder)
import DtvtLib
from DtvtLib import Utility, Configuration , CSVReader
param = Utility.Initialise(root_folder,__file__)
log_file_path    = param['log_file_path']
result_file_name = param['result_file_name']
result_file_path = param['result_file_path']
dtvt_log         = param['dtvt_log']
proj_const       = param['proj_const']
SyDT             = param['SyDT']
### ======== Template Ends==================================================================================================


### =========Rule RULE_DB_ROUTE_0003 ========================================================================================
# Author : Saleem Javed
# Updated 10 April 2017
# Caps Used : Signals_Cap ,Routes_Cap ,Switchs_Cap ,Points_Cap, Secondary_Detection_Devices_Cap,Signalisation_Areas_Cap
# Constant used :  None 
###========================================================================================================================== 

routes_cap = SyDT['Routes_Cap']
switchs_cap = SyDT['Switchs_Cap']
points_cap =  SyDT['Points_Cap']
sddevices_cap =SyDT['Secondary_Detection_Devices_Cap']
sig_area_cap = SyDT['Signalisation_Areas_Cap']
sdd_id_list = sig_area_cap['Area_Boundary_Secondary_Detection_Device_ID_List']
cbi_sig_area=dict()
test_result=''
# Create report file
global_test_results = list()
wbReport = workbook.Workbook()
wsRpt=wbReport.active
wsRpt.title = "Test_Results"
rwCount = 1
Utility.WriteToWorkSheet(wsRpt,rwCount,["Route", "Switchs", "Point", "SDD List", "Signalisation Areas","Result"])
rwCount = rwCount+1

for i in range(len(sig_area_cap['Name'])):
    if sig_area_cap['Area_Type'][i] =='CBI':
        cbi_sig_area[sig_area_cap['Name'][i]] = sdd_id_list[i].split(';')

route_names = routes_cap['Name']
switch_ids=routes_cap['Switch_ID_List']

for route in route_names:
    report=dict()
    sw_sig_area , sw_pt_list ,route_pt_list, sdd_list ,route_sdd_list= [],[],[] ,[],[]     
    route_switch_ids = switch_ids[route_names.index(route)] 

    if route_switch_ids != '0':

        for sw in  str.split(route_switch_ids,';'):
            sw_index = switchs_cap['Name'].index(sw)
            sw_pt_list = Utility.GetPointsOfSwitch(sw_index,switchs_cap)  
            route_pt_list.extend(sw_pt_list)
            sdd_name=''

            for pt in sw_pt_list:
                sdd_name= Utility.GetSDDofPoint(pt,sddevices_cap)
                route_sdd_list.append(sdd_name)

                if sdd_name == '':
                    dtvt_log("Error: Point " + pt + " is not linked with SDD")
                else:
                    for key,value in cbi_sig_area.items():
                        for vitem in value:
                            if sdd_name == vitem:                                
                                sw_sig_area.append(key)
                                break

        if len(set(sw_sig_area)) > 1:
            dtvt_log.error("For route " + route + " all switches do not belong to same CBI Signalisation Area")
            test_result='NOK'
            global_test_results.append(test_result)
        else:            
            dtvt_log.info("For route " + route + " all switches belong to same CBI Signalisation Area")            
            test_result = 'OK'
            global_test_results.append(test_result)

        wsRpt.cell(row = rwCount, column = 1, value = route)
        wsRpt.cell(row = rwCount, column = 2, value = route_switch_ids)
        wsRpt.cell(row = rwCount, column = 3, value = Utility.ConvertToString(route_pt_list))
        wsRpt.cell(row = rwCount, column = 4, value = Utility.ConvertToString( route_sdd_list))
        wsRpt.cell(row = rwCount, column = 5, value = Utility.ConvertToString(sw_sig_area))
        wsRpt.cell(row = rwCount, column = 6, value = test_result)
        Utility.WriteToWorkSheet(wsRpt,rwCount,[route, route_switch_ids, Utility.ConvertToString(route_pt_list) ,Utility.ConvertToString( route_sdd_list),Utility.ConvertToString(sw_sig_area),test_result])
        rwCount = rwCount+1



# Save Report
Utility.SaveReport(wbReport,global_test_results,root_folder,result_file_name)

 
    








