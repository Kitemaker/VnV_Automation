### ========= Template =========================================================================
import sys , os , os.path
import logging
from openpyxl import workbook, worksheet
# add root folder in search path
root_folder = os.path.abspath(os.path.join(os.path.dirname(__file__),".."))
sys.path.append(root_folder)
import DtvtLib
from DtvtLib import Utility, Configuration , CSVReader
param = Utility.Initialise(root_folder,__file__)
log_file_path    = param['log_file_path']
result_file_name = param['result_file_name']
result_file_path = param['result_file_path']
dtvt_log         = param['dtvt_log']
proj_const       = param['proj_const']
sydb_reader      =  param['SyDB']

### ======== Template Ends=================================================================================================

### =========Rule RULE_TMS_SIGNAL_0012 ====================================================================================
# Author : Saleem Javed
# Updated 13 April 2017
# Caps Used : SYDB Node used, Routes,Switch,Blocks,Points,Signals
# Constant used :  None
###========================================================================================================================
from DtvtLib import SydbReader
import xml.etree.ElementTree as ET

point_deadlocking_blocks = sydb_reader.Point_Deadlocking_Block_ID_List()
signal_kp_pair = sydb_reader.Get_Signal_ID_Kp_Pairs()
route_sig_pair = sydb_reader.Get_Route_OriginSignal_Pairs() 
point_id_name_dict = sydb_reader.Get_Point_ID_Name_Pairs()
sig_id_blk_dict =sydb_reader.Get_Signals_ID_BlockID_Pairs()


# Create Report
global_test_results = list()
wbReport = workbook.Workbook()
wsRpt=wbReport.active
wsRpt.title = "Test_Results"
rwCount = 1
wsRpt.cell(row = rwCount, column = 1, value = "Route")
wsRpt.cell(row = rwCount, column = 2, value = "Signal_Origin")
wsRpt.cell(row = rwCount, column = 3, value = "Signal_Origin_Block")
wsRpt.cell(row = rwCount, column = 4, value = "Points")
wsRpt.cell(row = rwCount, column = 5, value = "Route_Deadlocking_blocks")
wsRpt.cell(row = rwCount, column = 6, value = "Result")
rwCount = rwCount+1

for route in sydb_reader.root_node.findall('./Routes/Route'):
    route_name = route.attrib['Name']
    route_points = list()
    route_switchs = list()
    route_pt_deadlocking_blocks = list()
    result  = ""

    signal = route_sig_pair[route.attrib['Name']]   
    sig_kp = int(signal_kp_pair[signal])
    sig_blk = ''

    # Get block of signal
    sig_blk = sig_id_blk_dict[signal]

    # Get all points of the route
    points_of_routes = sydb_reader.Get_Route_Points_Pairs()

    #sw_pts_pair = sydb_reader   
    route_deadlocking_blocks = list()

    # Test the Rule
    for pt in points_of_routes[route_name]:

        # get deadlocking blocks of the point and add them in the rout deadlocking blocks
        route_deadlocking_blocks.extend( point_deadlocking_blocks[point_id_name_dict[pt]]) 
                       
        # check if signal block is in the deadlocking blocks of the point if yes break 
        if sig_blk in point_deadlocking_blocks[point_id_name_dict[pt]]:
            dtvt_log.error( "For Route : " + route_name + " Signal " + signal + " with block ID = " + sig_blk + "in deadlocking blocks of point " + point_id_name_dict[pt])
            result = "NOK"
            break
        
        result = "OK"

    # Save the result for the route in excel
    global_test_results.append(result)
    wsRpt.cell(row = rwCount, column = 1, value = route_name)
    wsRpt.cell(row = rwCount, column = 2, value = signal)
    wsRpt.cell(row = rwCount, column = 3, value = sig_blk)
    wsRpt.cell(row = rwCount, column = 4, value = Utility.ConvertToString(points_of_routes[route_name]))
    wsRpt.cell(row = rwCount, column = 5, value = Utility.ConvertToString(route_deadlocking_blocks))
    wsRpt.cell(row = rwCount, column = 6, value = result)
    rwCount = rwCount+1

# Save Report
Utility.SaveReport(wbReport,global_test_results,root_folder,result_file_name)







   
        

    
    





    
