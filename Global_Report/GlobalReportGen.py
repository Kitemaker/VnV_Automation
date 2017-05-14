import sys
from openpyxl import workbook
from openpyxl import worksheet
from openpyxl import load_workbook
from openpyxl import cell
import GlobalReport_Constant as GConst
import Utility

fPathInput=r"C:\Users\295563\Documents\01 PROJECTS\V&V\PANAMA\GlobalReport_Gen_Panama_vC1.xlsm"
fPathConsolidatedReport=r"C:\Users\295563\Documents\01 PROJECTS\V&V\PANAMA\Panama IVR_Status tracker_ver_2A.xlsm"
print("Running sig rule for :"+ fPathInput)
try:
    inputWb=load_workbook(fPathInput,keep_vba=True)
except FileNotFoundError:
    print("File Not found, Please check the path")
    exit() 

#worksheet for full list of rule from UEVOL0962    
wsUevol0962=inputWb.get_sheet_by_name(GConst.c_WsNameUevol0962)
# worksheet for design applicability of GfPD rule
wsTraceability =inputWb.get_sheet_by_name(GConst.c_WsNameTraceability)
# worksheet for Auto -Manu
wsAutoManu= inputWb.get_sheet_by_name(GConst.c_WsNameAutoManu)
# worksheet for safety and grouping of GfPD rule
wsSRSS =inputWb.get_sheet_by_name(GConst.c_WsNameSRSS)
# worksheet for Auto Rule Results
wsDoubleCheck =inputWb.get_sheet_by_name(GConst.c_WsNameAutoResults)
# worksheet for design applicability and safety of Specific rules
wsSpecific =inputWb.get_sheet_by_name(GConst.c_WsNameSpecific)

fout=open('out.txt','w')
#fconsolidatedReport =open('ConsolidatedReport.txt','w')

rule_Uevol=[]
rule_Specific=[]
rule_Name_List=[]
rule_Name_List_Applica=[]
rule_Name_List_Safety=[]
rule_Name_List_Auto=[]

rule_Name_Applicability=[]
rule_Applicability=[]
rule_Name_Safety=[]
rule_Safety=[]
rule_Name_AutoManu=[]
rule_Auto=[]

temp = list(wsUevol0962[GConst.c_ColUevolRuleName])
for i in range(1,len(temp)):
    rule_Uevol.append(temp[i].value)

temp=list(wsSpecific[GConst.c_ColSpecRuleName])
for i in range(1,len(temp)):
    rule_Specific.append(temp[i].value)

rule_Name_List=rule_Uevol+rule_Specific


temp =list(wsTraceability[GConst.c_ColTraceRule]) +list(wsSpecific[GConst.c_ColSpecRuleName])
for i in range(len(temp)):
    rule_Name_Applicability.append(temp[i].value)

temp =list(wsTraceability[GConst.c_ColTraceDesApp])+ list(wsSpecific[GConst.c_ColSpecAppl])
for i in range(len(temp)):
    rule_Applicability.append(temp[i].value)

temp =  list(wsSRSS[GConst.c_ColSSRSRuleName])+list(wsSpecific[GConst.c_ColSpecRuleName])
for i in range(len(temp)):
    rule_Name_Safety.append(temp[i].value)

temp= list(wsSRSS[GConst.c_ColSSRSRuleSafety]) + list(wsSpecific[GConst.c_ColSpecSafety])
for i in range(len(temp)):
    rule_Safety.append(temp[i].value)

temp = list(wsAutoManu[GConst.c_ColAutoManRuleName])
for i in range(len(temp)):
    rule_Name_AutoManu.append(temp[i].value)
temp = list( wsAutoManu[GConst.c_ColAutoMan])
for i in range(len(temp)):
    rule_Auto.append(temp[i].value)

#Release worksheets

del(wsUevol0962)
del(wsTraceability)
del(wsAutoManu)
del(wsSRSS)
del(wsDoubleCheck)
del(wsSpecific) 



ruleForm="{:<30}  ,  {:20}  ,{:20}  ,{:20}  "

for name in rule_Name_List:
    # Get applicability
    if (Utility.isItemInList(name,rule_Name_Applicability)):
        rule_Name_List_Applica.append(rule_Applicability[rule_Name_Applicability.index(name)])
    else:
        rule_Name_List_Applica.append("Not Found")
    # Get Safety
    if (Utility.isItemInList(name,rule_Name_Safety)):
        rule_Name_List_Safety.append(rule_Safety[rule_Name_Safety.index(name)])
    else:
        rule_Name_List_Safety.append("Not Found")
    # Get Auto/Manu
    if (Utility.isItemInList(name,rule_Name_AutoManu)):
        rule_Name_List_Auto.append(rule_Auto[rule_Name_AutoManu.index(name)])
    else:
        rule_Name_List_Auto.append("Not Found")



for i in range(len(rule_Name_List)):
    try:
        data=ruleForm.format(rule_Name_List[i],rule_Name_List_Applica[i],rule_Name_List_Safety[i],rule_Name_List_Auto[i])
        dataforFile="{},{},{},{}".format(rule_Name_List[i],rule_Name_List_Applica[i],rule_Name_List_Safety[i],rule_Name_List_Auto[i])
       
        fout.write(dataforFile+"\n")
    except:
        print("Unexpected error:Rule Name: ",i,"\t", sys.exc_info()[0])
        #break


try:
    consolReportWb=load_workbook(fPathConsolidatedReport,keep_vba=True)
except FileNotFoundError:
    print("File Not found, Please check the path for" + fPathConsolidatedReport)
    exit()




#worksheet for full list of rule from UEVOL0962    
wsConsolReport=consolReportWb.get_sheet_by_name(GConst.c_WsNameConsReport)

colRuleName= Utility.ReplaceNoneCell(list(wsConsolReport['A']))
colSafety=Utility.ReplaceNoneCell(list(wsConsolReport['B']))
colTypeOfRule=Utility.ReplaceNoneCell(list(wsConsolReport['C']))
colCh1Pic=Utility.ReplaceNoneCell(list(wsConsolReport['D']))
colCh1Res=Utility.ReplaceNoneCell(list(wsConsolReport['E']))
colCh1Com=Utility.ReplaceNoneCell(list(wsConsolReport['F']))
colCh2Pic=Utility.ReplaceNoneCell(list(wsConsolReport['H']))
colCh2Res=Utility.ReplaceNoneCell(list(wsConsolReport['I']))
colCh2Com=Utility.ReplaceNoneCell(list(wsConsolReport['J']))
colTotRes=Utility.ReplaceNoneCell(list(wsConsolReport['L']))
colTotCom=Utility.ReplaceNoneCell(list(wsConsolReport['M']))
colRuleType=Utility.ReplaceNoneCell(list(wsConsolReport['O']))



manualRule_GFPD=[]
manualRule_SyCR=[]
manualRule_SyDB=[]



for i in range( len(colRuleName)):
    # Get Result Type
    try:        
        sout= colRuleName[i],colSafety[i],colCh1Pic[i],colCh1Res[i],colCh1Com[i],colCh2Pic[i],colCh2Res[i],colCh2Com[i],colTotRes[i],colTotCom[i]
        
        if (colRuleType[i] =="SyDT"):
            if(colTypeOfRule[i]=="GFPD"):
                manualRule_GFPD.append(list(sout))
            if(colTypeOfRule[i]=="SyCR"):
                manualRule_SyCR.append(list(sout))
        elif(colRuleType[i] =="SyDB"):
            manualRule_SyDB.append(list(sout))
        else:
            print("Unexpected error:Rule Type  not found in consolidated report: ","\t", colRuleName[i])
        
       # fconsolidatedReport.write(colRuleName[i]+","+colSafety[i]+","+colCh1Pic[i]+","+colCh1Res[i]+","+colCh1Com[i]+","+colCh2Pic[i]+","+colCh2Res[i]+","+colCh2Com[i]+","+colTotRes[i]+","+colTotCom[i]+","+"\n")
    except:
        print("Unexpected erro in reading consolidated result :Rule number: ",i,"\t", sys.exc_info()[0])
 
wbTemp = load_workbook(r"C:\Users\295563\Documents\01 PROJECTS\V&V\PANAMA\PNML1C_Global Report_SyDT_VR_ver1A.xlsx")
wsManualRuleGfpd=wbTemp.get_sheet_by_name(GConst.c_WsNameGfpd)
wsManualRuleSyDB=wbTemp.get_sheet_by_name(GConst.c_WsNameSyDB)
wsManualRuleSyCR=wbTemp.get_sheet_by_name(GConst.c_WsNameSyCR)

rg=wsManualRuleGfpd['A:Z']
for col in rg:
    for j in range(3,len(col)):
        col[j].value=""

rg=wsManualRuleSyDB['A:Z']
for col in rg:
    for j in range(3,len(col)):
        col[j].value=""
rg=wsManualRuleSyCR['A:Z']
for col in rg:
    for j in range(3,len(col)):
        col[j].value=""

j=4
for ritem in range(len(manualRule_GFPD)):
    item=manualRule_GFPD[ritem]
    try:
        wsManualRuleGfpd['A'+str(j)]=str(ritem)
        wsManualRuleGfpd['B'+str(j)]=item[0]
        wsManualRuleGfpd['C'+str(j)]=item[1]
        wsManualRuleGfpd['D'+str(j)]=item[2]
        wsManualRuleGfpd['E'+str(j)]=item[3]
        wsManualRuleGfpd['F'+str(j)]=item[4]
        wsManualRuleGfpd['G'+str(j)]=item[5]
        wsManualRuleGfpd['H'+str(j)]=item[6]
        wsManualRuleGfpd['I'+str(j)]=item[7]
        wsManualRuleGfpd['J'+str(j)]=item[8]
        wsManualRuleGfpd['K'+str(j)]=item[9]
        j=j+1
    except:
        print("Unexpected error range(len(manualRule_GFPD)):"+"\t"+sys.exc_info()[0])

j=4
for ritem1 in range(len(manualRule_SyDB)):
    item=manualRule_SyDB[ritem1]
    try:
        wsManualRuleSyDB['A'+str(j)]=str(ritem1)
        wsManualRuleSyDB['B'+str(j)]=item[0]
        wsManualRuleSyDB['C'+str(j)]=item[1]
        wsManualRuleSyDB['D'+str(j)]=item[2]
        wsManualRuleSyDB['E'+str(j)]=item[3]
        wsManualRuleSyDB['F'+str(j)]=item[4]
        wsManualRuleSyDB['G'+str(j)]=item[5]
        wsManualRuleSyDB['H'+str(j)]=item[6]
        wsManualRuleSyDB['I'+str(j)]=item[7]
        wsManualRuleSyDB['J'+str(j)]=item[8]
        wsManualRuleSyDB['K'+str(j)]=item[9]
        j=j+1
    except:
        print("Unexpected error range(len(wsManualRuleSyDB)):"+"\t"+sys.exc_info()[0])

j=4
for ritem2 in range(len(manualRule_SyCR)):
    item=manualRule_SyCR[ritem2]
    try:
        wsManualRuleSyCR['A'+str(j)]=str(ritem2)
        wsManualRuleSyCR['B'+str(j)]=item[0]
        wsManualRuleSyCR['C'+str(j)]=item[1]
        wsManualRuleSyCR['D'+str(j)]=item[2]
        wsManualRuleSyCR['E'+str(j)]=item[3]
        wsManualRuleSyCR['F'+str(j)]=item[4]
        wsManualRuleSyCR['G'+str(j)]=item[5]
        wsManualRuleSyCR['H'+str(j)]=item[6]
        wsManualRuleSyCR['I'+str(j)]=item[7]
        wsManualRuleSyCR['J'+str(j)]=item[8]
        wsManualRuleSyCR['K'+str(j)]=item[9]
        j=j+1
    except:
        print("Unexpected error range(len(wsManualRuleSyDB)):"+"\t"+sys.exc_info()[0])

wbTemp.save(r"C:\Users\295563\Documents\01 PROJECTS\V&V\PANAMA\PNML1C_Global Report_SyDT_VR_ver1A.xlsx")